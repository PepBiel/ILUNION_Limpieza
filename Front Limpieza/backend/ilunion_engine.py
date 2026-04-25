from __future__ import annotations

import calendar
import datetime as dt
import math
import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import scipy.sparse as sp
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp


SHIFT_HOURS = {"M": 7, "T": 7, "N": 10}
WORK_CODES = ["M", "T", "N"]
REST_CODE = "D"

WEEKDAY_NAMES = {
    0: "LUNES",
    1: "MARTES",
    2: "MIERCOLES",
    3: "JUEVES",
    4: "VIERNES",
    5: "SABADO",
    6: "DOMINGO",
}

SUMMER_MONTHS = {6, 7, 8}
SUMMER_MIN_REST = 10


@dataclass
class Config:
    hospital: str
    year: int
    workers_file: Path
    presences_file: Path
    output_file: Path
    holidays_file: Optional[Path] = None
    fixed_assignments_file: Optional[Path] = None
    time_limit_per_month: int = 30

    understaff_penalty: float = 150.0
    over_hours_penalty: float = 3.0
    under_hours_penalty: float = 3.0
    deviation_penalty: float = 0.2
    off_day_penalty: float = 0.1


@dataclass
class AnnualRunArtifacts:
    config: Config
    workers: pd.DataFrame
    demands: Dict[str, Dict[Tuple[str, str], int]]
    holidays: set[dt.date]
    fixed_assignments: Dict[Tuple[str, dt.date], str]
    monthly_schedules: Dict[int, pd.DataFrame]
    coverage_all: pd.DataFrame
    monthly_hours_all: pd.DataFrame
    annual_hours: pd.DataFrame
    solve_log: pd.DataFrame


def normalize_text(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip().upper()


def normalize_key(value: object) -> str:
    text = normalize_text(value)
    text = unicodedata.normalize("NFKD", text)
    return text.encode("ascii", "ignore").decode("ascii")


def hospital_sheet_name(hospital: str) -> str:
    hospital_key = normalize_key(hospital)
    if "CLINICO" in hospital_key:
        return "H.CLINICO"
    if "GIL" in hospital_key:
        return "H.GIL CASARES"
    raise ValueError(f"Hospital no reconocido: {hospital}")


def worker_matches_hospital(center_value: object, hospital: str) -> bool:
    center = normalize_key(center_value)
    hospital_key = normalize_key(hospital)
    if "CLINICO" in hospital_key:
        return "CLINICO" in center
    if "GIL" in hospital_key:
        return "GIL" in center
    return False


def infer_category(row: pd.Series) -> str:
    puesto = normalize_key(row.get("PUESTO", ""))
    return "PEON" if "PEON" in puesto else "LIMPIADOR"


def infer_allowed_shifts(row: pd.Series) -> List[str]:
    turno = normalize_key(row.get("TURNO", ""))
    observations = normalize_key(row.get("observaciones", ""))
    category = row["categoria"]

    if "NOCHE/TARDE" in turno or "NOCHES Y TARDE" in observations:
        return ["N", "T"]
    if "NOCHE" in turno or "UNA NOCHE SI UNA NOCHE NO" in observations:
        return ["N"]
    if "CONCILIACION MANANA" in turno or "CONCILIACION" in observations:
        return ["M"]
    if "DE LUNES A VIERNES" in observations:
        return ["M"] if category != "PEON" else ["M", "T"]
    if (
        "15 DIAS MANANA/15 DIAS TARDE" in observations
        or "1 MES DE MANANA/1 MES DE TARDE" in observations
    ):
        return ["M", "T"]
    if "ROTATORIO" in turno:
        return ["M", "T"]
    if turno == "MANANA":
        return ["M"]
    if turno == "TARDE":
        return ["T"]
    if category == "PEON":
        return ["M", "T"]
    return ["M", "T"]


def preferred_shift(row: pd.Series) -> Optional[str]:
    turno = normalize_key(row.get("TURNO", ""))
    if "NOCHE" in turno:
        return "N"
    if "TARDE" in turno:
        return "T"
    if "MANANA" in turno or "CONCILIACION" in turno:
        return "M"
    return None


def parse_weekend_ratio(row: pd.Series) -> Optional[int]:
    observations = normalize_key(row.get("observaciones", ""))
    if "UN FIN DE SEMANA SI Y DOS" in observations:
        return 3
    if (
        "UN FIN DE SEMANA SI Y UNO NO" in observations
        or "UN SABADO SI Y UNO NO" in observations
    ):
        return 2
    return None


def load_holidays(path: Optional[Path]) -> set[dt.date]:
    if path is None:
        return set()

    dataframe = pd.read_csv(path) if path.suffix.lower() == ".csv" else pd.read_excel(path)
    if "date" not in dataframe.columns:
        dataframe = dataframe.rename(columns={dataframe.columns[0]: "date"})

    return {
        pd.to_datetime(value).date()
        for value in dataframe["date"]
        if not pd.isna(value)
    }


def load_fixed_assignments(path: Optional[Path]) -> Dict[Tuple[str, dt.date], str]:
    if path is None:
        return {}

    dataframe = pd.read_csv(path) if path.suffix.lower() == ".csv" else pd.read_excel(path)
    columns = {normalize_key(column): column for column in dataframe.columns}
    required = ["TRABAJADOR", "DATE", "CODE"]
    missing = [column for column in required if column not in columns]
    if missing:
        raise ValueError(f"Faltan columnas en fixed_assignments: {missing}")

    assignments: Dict[Tuple[str, dt.date], str] = {}
    for _, row in dataframe.iterrows():
        worker = str(row[columns["TRABAJADOR"]]).strip()
        date = pd.to_datetime(row[columns["DATE"]]).date()
        code = normalize_text(row[columns["CODE"]])
        assignments[(worker, date)] = code

    return assignments


def load_workers(path: Path, hospital: str) -> pd.DataFrame:
    workers = pd.read_excel(path)
    workers = workers[
        workers["CENTRO"].apply(lambda value: worker_matches_hospital(value, hospital))
    ].copy()
    if workers.empty:
        raise ValueError(f"No se encontraron trabajadores para {hospital}")

    if "observaciones" not in workers.columns:
        workers["observaciones"] = ""

    workers["TRABAJADOR"] = workers["TRABAJADOR"].astype(str).str.strip()
    workers["categoria"] = workers.apply(infer_category, axis=1)
    workers["allowed_shifts"] = workers.apply(infer_allowed_shifts, axis=1)
    workers["preferred_shift"] = workers.apply(preferred_shift, axis=1)
    workers["weekend_ratio"] = workers.apply(parse_weekend_ratio, axis=1)
    workers["annual_hours"] = pd.to_numeric(
        workers["HORAS/A\u00d1O"], errors="coerce"
    ).fillna(1715.0)
    workers["obs_norm"] = workers["observaciones"].apply(normalize_key)
    return workers.reset_index(drop=True)


def load_demands(path: Path, hospital: str) -> Dict[str, Dict[Tuple[str, str], int]]:
    sheet_name = hospital_sheet_name(hospital)
    dataframe = pd.read_excel(path, sheet_name=sheet_name, header=None)

    day_names = [
        normalize_key(value)
        for value in dataframe.iloc[1, [0, 5, 10, 15, 20, 25, 30, 35]].tolist()
    ]
    values = dataframe.iloc[3].tolist()

    demands: Dict[str, Dict[Tuple[str, str], int]] = {}
    for block_index, day_name in enumerate(day_names):
        base = block_index * 5
        demands[day_name] = {
            ("LIMPIADOR", "M"): int(math.ceil(float(values[base] or 0))),
            ("LIMPIADOR", "T"): int(math.ceil(float(values[base + 1] or 0))),
            ("LIMPIADOR", "N"): int(math.ceil(float(values[base + 2] or 0))),
            ("PEON", "M"): int(math.ceil(float(values[base + 3] or 0))),
            ("PEON", "T"): int(math.ceil(float(values[base + 4] or 0))),
        }
    return demands


class MonthlySolver:
    def __init__(
        self,
        workers: pd.DataFrame,
        dates: List[dt.date],
        demands: Dict[str, Dict[Tuple[str, str], int]],
        holidays: set[dt.date],
        fixed_assignments: Dict[Tuple[str, dt.date], str],
        config: Config,
        annual_hours_remaining: Dict[int, float],
        months_remaining_including_current: int,
        previous_last_code: Dict[int, Optional[str]],
    ):
        self.workers = workers
        self.dates = dates
        self.demands = demands
        self.holidays = holidays
        self.fixed_assignments = fixed_assignments
        self.config = config
        self.annual_hours_remaining = annual_hours_remaining
        self.months_remaining = months_remaining_including_current
        self.previous_last_code = previous_last_code

        self.var_index: Dict[Tuple[object, ...], int] = {}
        self.obj: List[float] = []
        self.integrality: List[int] = []
        self.lb: List[float] = []
        self.ub: List[float] = []
        self.rows: List[Tuple[List[int], List[float]]] = []
        self.lo: List[float] = []
        self.hi: List[float] = []

    def add_var(
        self,
        key: Tuple[object, ...],
        cost: float,
        integer: bool = True,
        lower: float = 0.0,
        upper: float = 1.0,
    ) -> int:
        index = len(self.obj)
        self.var_index[key] = index
        self.obj.append(cost)
        self.integrality.append(1 if integer else 0)
        self.lb.append(lower)
        self.ub.append(upper)
        return index

    def add_constraint(
        self,
        items: Iterable[Tuple[int, float]],
        lower: float,
        upper: float,
    ) -> None:
        indexes, coefficients = zip(*items) if items else ([], [])
        self.rows.append((list(indexes), list(coefficients)))
        self.lo.append(lower)
        self.hi.append(upper)

    def day_type(self, date: dt.date) -> str:
        if date in self.holidays:
            return "FESTIVOS"
        return WEEKDAY_NAMES[date.weekday()]

    def target_hours_for_worker(self, worker_index: int) -> float:
        remaining = max(0.0, self.annual_hours_remaining[worker_index])
        return remaining / max(1, self.months_remaining)

    def build(self) -> None:
        for worker_index, row in self.workers.iterrows():
            allowed_shifts = list(row["allowed_shifts"])
            preferred = row["preferred_shift"]

            for day_index, date in enumerate(self.dates):
                fixed_code = self.fixed_assignments.get((row["TRABAJADOR"], date))
                for code in allowed_shifts + [REST_CODE]:
                    lower, upper = 0.0, 1.0
                    if fixed_code is not None:
                        lower = upper = 1.0 if code == fixed_code else 0.0

                    cost = self.config.off_day_penalty if code == REST_CODE else 0.0
                    if preferred and code in WORK_CODES and code != preferred:
                        cost += self.config.deviation_penalty

                    self.add_var(
                        ("x", worker_index, day_index, code),
                        cost,
                        integer=True,
                        lower=lower,
                        upper=upper,
                    )

        for day_index, date in enumerate(self.dates):
            day_name = self.day_type(date)
            for category in ["LIMPIADOR", "PEON"]:
                for code in WORK_CODES:
                    need = self.demands.get(day_name, {}).get((category, code), 0)
                    if need > 0:
                        self.add_var(
                            ("u", day_index, category, code),
                            self.config.understaff_penalty,
                            integer=False,
                            lower=0.0,
                            upper=float(need),
                        )

        for worker_index, _ in self.workers.iterrows():
            self.add_var(
                ("over", worker_index),
                self.config.over_hours_penalty,
                integer=False,
                lower=0.0,
                upper=np.inf,
            )
            self.add_var(
                ("under", worker_index),
                self.config.under_hours_penalty,
                integer=False,
                lower=0.0,
                upper=np.inf,
            )

        for worker_index, row in self.workers.iterrows():
            allowed_shifts = list(row["allowed_shifts"])
            for day_index, _ in enumerate(self.dates):
                items = [
                    (self.var_index[("x", worker_index, day_index, code)], 1.0)
                    for code in allowed_shifts + [REST_CODE]
                ]
                self.add_constraint(items, 1.0, 1.0)

        for day_index, date in enumerate(self.dates):
            day_name = self.day_type(date)
            for category in ["LIMPIADOR", "PEON"]:
                for code in WORK_CODES:
                    need = self.demands.get(day_name, {}).get((category, code), 0)
                    if need <= 0:
                        continue

                    items = []
                    for worker_index, row in self.workers.iterrows():
                        if row["categoria"] == category and code in row["allowed_shifts"]:
                            items.append(
                                (self.var_index[("x", worker_index, day_index, code)], 1.0)
                            )

                    items.append((self.var_index[("u", day_index, category, code)], 1.0))
                    self.add_constraint(items, float(need), np.inf)

        for worker_index, row in self.workers.iterrows():
            target = self.target_hours_for_worker(worker_index)
            items = []
            for day_index, _ in enumerate(self.dates):
                for code in row["allowed_shifts"]:
                    items.append(
                        (
                            self.var_index[("x", worker_index, day_index, code)],
                            float(SHIFT_HOURS[code]),
                        )
                    )

            items.append((self.var_index[("over", worker_index)], -1.0))
            items.append((self.var_index[("under", worker_index)], 1.0))
            self.add_constraint(items, target, target)

        for worker_index, row in self.workers.iterrows():
            if "DE LUNES A VIERNES" not in row["obs_norm"]:
                continue

            for day_index, date in enumerate(self.dates):
                if date.weekday() >= 5:
                    for code in row["allowed_shifts"]:
                        self.add_constraint(
                            [(self.var_index[("x", worker_index, day_index, code)], 1.0)],
                            0.0,
                            0.0,
                        )

        for worker_index, row in self.workers.iterrows():
            if "N" not in row["allowed_shifts"]:
                continue

            if self.previous_last_code.get(worker_index) == "N" and self.dates:
                self.add_constraint(
                    [(self.var_index[("x", worker_index, 0, "N")], 1.0)],
                    0.0,
                    0.0,
                )

            for day_index in range(len(self.dates) - 1):
                self.add_constraint(
                    [
                        (self.var_index[("x", worker_index, day_index, "N")], 1.0),
                        (self.var_index[("x", worker_index, day_index + 1, "N")], 1.0),
                    ],
                    -np.inf,
                    1.0,
                )

        weekend_days = [
            index for index, date in enumerate(self.dates) if date.weekday() >= 5
        ]
        weekends_in_month = len(
            {
                (date.isocalendar().year, date.isocalendar().week)
                for date in self.dates
                if date.weekday() >= 5
            }
        )
        for worker_index, row in self.workers.iterrows():
            ratio = row["weekend_ratio"]
            if pd.isna(ratio) or not ratio or not weekend_days:
                continue

            max_weekend_days = math.ceil((weekends_in_month / ratio) * 2.0)
            items = []
            for day_index in weekend_days:
                for code in row["allowed_shifts"]:
                    items.append(
                        (self.var_index[("x", worker_index, day_index, code)], 1.0)
                    )
            self.add_constraint(items, -np.inf, float(max_weekend_days))

        month = self.dates[0].month
        if month in SUMMER_MONTHS:
            for worker_index, _ in self.workers.iterrows():
                items = [
                    (self.var_index[("x", worker_index, day_index, REST_CODE)], 1.0)
                    for day_index, _ in enumerate(self.dates)
                ]
                self.add_constraint(items, float(SUMMER_MIN_REST), np.inf)

    def solve(self) -> Tuple[pd.DataFrame, object]:
        self.build()

        variable_count = len(self.obj)
        row_count = len(self.rows)
        data, row_indexes, column_indexes = [], [], []

        for row_index, (indexes, coefficients) in enumerate(self.rows):
            data.extend(coefficients)
            row_indexes.extend([row_index] * len(indexes))
            column_indexes.extend(indexes)

        matrix = sp.coo_array(
            (data, (row_indexes, column_indexes)),
            shape=(row_count, variable_count),
        ).tocsr()

        result = milp(
            c=np.array(self.obj),
            integrality=np.array(self.integrality),
            bounds=Bounds(self.lb, self.ub),
            constraints=LinearConstraint(matrix, self.lo, self.hi),
            options={"time_limit": self.config.time_limit_per_month},
        )

        if result.x is None:
            raise RuntimeError(f"No se obtuvo solucion: {result.message}")

        return self.extract_solution(result.x), result

    def extract_solution(self, values: np.ndarray) -> pd.DataFrame:
        rows = []
        for worker_index, row in self.workers.iterrows():
            output = {
                "WORKER_IDX": worker_index,
                "TRABAJADOR": row["TRABAJADOR"],
                "CENTRO": row["CENTRO"],
                "TURNO_BASE": row["TURNO"],
                "PUESTO": row["PUESTO"],
                "CATEGORIA": row["categoria"],
                "HORAS_A\u00d1O": row["annual_hours"],
            }
            for day_index, date in enumerate(self.dates):
                chosen = REST_CODE
                for code in list(row["allowed_shifts"]) + [REST_CODE]:
                    variable_index = self.var_index[("x", worker_index, day_index, code)]
                    if values[variable_index] > 0.5:
                        chosen = code
                        break
                output[date.isoformat()] = chosen
            rows.append(output)

        return pd.DataFrame(rows)


def build_coverage(
    schedule: pd.DataFrame,
    demands: Dict[str, Dict[Tuple[str, str], int]],
    holidays: set[dt.date],
    dates: List[dt.date],
) -> pd.DataFrame:
    records = []
    for date in dates:
        date_column = date.isoformat()
        day_type = "FESTIVOS" if date in holidays else WEEKDAY_NAMES[date.weekday()]
        for category in ["LIMPIADOR", "PEON"]:
            for code in WORK_CODES:
                need = demands.get(day_type, {}).get((category, code), 0)
                if need <= 0:
                    continue

                assigned = int(
                    ((schedule["CATEGORIA"] == category) & (schedule[date_column] == code)).sum()
                )
                records.append(
                    {
                        "fecha": date,
                        "tipo_dia": day_type,
                        "categoria": category,
                        "turno": code,
                        "necesario": need,
                        "asignado": assigned,
                        "desviacion": assigned - need,
                        "cumple": assigned >= need,
                    }
                )

    return pd.DataFrame(records)


def build_month_hours(
    schedule: pd.DataFrame,
    target_hours_by_worker: Dict[int, float],
    dates: List[dt.date],
) -> pd.DataFrame:
    records = []
    for _, row in schedule.iterrows():
        worker_index = int(row["WORKER_IDX"])
        assigned_hours = 0
        worked_shifts = 0
        rests = 0

        for date in dates:
            code = row[date.isoformat()]
            if code in SHIFT_HOURS:
                assigned_hours += SHIFT_HOURS[code]
                worked_shifts += 1
            elif code == REST_CODE:
                rests += 1

        records.append(
            {
                "TRABAJADOR": row["TRABAJADOR"],
                "CATEGORIA": row["CATEGORIA"],
                "HORAS_OBJETIVO_MES": round(target_hours_by_worker[worker_index], 2),
                "HORAS_ASIGNADAS_MES": assigned_hours,
                "DESVIACION_HORAS": round(
                    assigned_hours - target_hours_by_worker[worker_index], 2
                ),
                "TURNOS_TRABAJADOS": worked_shifts,
                "DESCANSOS": rests,
            }
        )

    return pd.DataFrame(records)


def write_annual_workbook(
    monthly_schedules: Dict[int, pd.DataFrame],
    coverage_all: pd.DataFrame,
    monthly_hours_all: pd.DataFrame,
    annual_hours: pd.DataFrame,
    solve_log: pd.DataFrame,
    config: Config,
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    weekend_fill = PatternFill("solid", fgColor="FFF2CC")
    summer_fill = PatternFill("solid", fgColor="E2F0D9")
    meta_fill = PatternFill("solid", fgColor="D9EAF7")

    for month, schedule in monthly_schedules.items():
        worksheet = workbook.create_sheet(
            f"{month:02d}_{calendar.month_name[month][:3].upper()}"
        )

        fixed_columns = [
            "MES",
            "TRABAJADOR",
            "CENTRO",
            "TURNO_BASE",
            "PUESTO",
            "CATEGORIA",
            "HORAS_A\u00d1O",
        ]
        date_columns = [
            column
            for column in schedule.columns
            if isinstance(column, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", column)
        ]
        columns = fixed_columns + date_columns

        for column_index, column in enumerate(columns, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_index, row in enumerate(
            schedule[columns].itertuples(index=False),
            start=2,
        ):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        for column_index, column in enumerate(
            date_columns,
            start=len(fixed_columns) + 1,
        ):
            date = dt.date.fromisoformat(column)
            if date.weekday() >= 5:
                for row_index in range(2, worksheet.max_row + 1):
                    worksheet.cell(row_index, column_index).fill = weekend_fill
            if month in SUMMER_MONTHS:
                for row_index in range(2, worksheet.max_row + 1):
                    if worksheet.cell(row_index, column_index).value == REST_CODE:
                        worksheet.cell(row_index, column_index).fill = summer_fill

        widths = {"A": 8, "B": 30, "C": 14, "D": 16, "E": 18, "F": 14, "G": 12}
        for column_letter, width in widths.items():
            worksheet.column_dimensions[column_letter].width = width
        for column_index in range(8, 8 + len(date_columns)):
            worksheet.column_dimensions[get_column_letter(column_index)].width = 12
        worksheet.freeze_panes = "H2"

    coverage_sheet = workbook.create_sheet("COBERTURA_ANUAL")
    for column_index, column in enumerate(coverage_all.columns, start=1):
        cell = coverage_sheet.cell(row=1, column=column_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(coverage_all.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            coverage_sheet.cell(row=row_index, column=column_index, value=value)
    for column_index in range(1, coverage_sheet.max_column + 1):
        coverage_sheet.column_dimensions[get_column_letter(column_index)].width = 16
    coverage_sheet.freeze_panes = "A2"

    monthly_hours_sheet = workbook.create_sheet("HORAS_MENSUALES")
    for column_index, column in enumerate(monthly_hours_all.columns, start=1):
        cell = monthly_hours_sheet.cell(row=1, column=column_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(monthly_hours_all.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            monthly_hours_sheet.cell(row=row_index, column=column_index, value=value)
    for column_index in range(1, monthly_hours_sheet.max_column + 1):
        monthly_hours_sheet.column_dimensions[get_column_letter(column_index)].width = 20
    monthly_hours_sheet.freeze_panes = "A2"

    annual_hours_sheet = workbook.create_sheet("HORAS_ANUALES")
    for column_index, column in enumerate(annual_hours.columns, start=1):
        cell = annual_hours_sheet.cell(row=1, column=column_index, value=column)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(annual_hours.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            annual_hours_sheet.cell(row=row_index, column=column_index, value=value)
    for column_index in range(1, annual_hours_sheet.max_column + 1):
        annual_hours_sheet.column_dimensions[get_column_letter(column_index)].width = 22
    annual_hours_sheet.freeze_panes = "A2"

    summary_sheet = workbook.create_sheet("RESUMEN")
    summary = [
        ("hospital", config.hospital),
        ("a\u00f1o", config.year),
        (
            "regla_verano",
            f"minimo {SUMMER_MIN_REST} descansos en junio/julio/agosto",
        ),
        ("faltas_cobertura_totales", int((~coverage_all["cumple"]).sum())),
        (
            "desviacion_media_anual_horas",
            float(annual_hours["DESVIACION_A\u00d1O"].mean()),
        ),
    ]
    for row_index, (key, value) in enumerate(summary, start=1):
        summary_sheet.cell(row_index, 1, key).fill = meta_fill
        summary_sheet.cell(row_index, 2, value)

    start_row = len(summary) + 3
    for column_index, column in enumerate(solve_log.columns, start=1):
        cell = summary_sheet.cell(start_row, column_index, column)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(
        solve_log.itertuples(index=False),
        start=start_row + 1,
    ):
        for column_index, value in enumerate(row, start=1):
            summary_sheet.cell(row_index, column_index, value)
    for column_index in range(1, max(3, summary_sheet.max_column) + 1):
        summary_sheet.column_dimensions[get_column_letter(column_index)].width = 24

    config.output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(config.output_file)


def run(config: Config) -> AnnualRunArtifacts:
    workers = load_workers(config.workers_file, config.hospital)
    demands = load_demands(config.presences_file, config.hospital)
    holidays = load_holidays(config.holidays_file)
    fixed_assignments = load_fixed_assignments(config.fixed_assignments_file)

    annual_remaining = {
        int(worker_index): float(value)
        for worker_index, value in workers["annual_hours"].items()
    }
    previous_last_code: Dict[int, Optional[str]] = {
        int(worker_index): None for worker_index in workers.index
    }

    monthly_schedules: Dict[int, pd.DataFrame] = {}
    coverage_list = []
    month_hours_list = []
    solve_log_rows = []

    for month in range(1, 13):
        dates = [
            dt.date(config.year, month, day)
            for day in range(1, calendar.monthrange(config.year, month)[1] + 1)
        ]
        target_hours_by_worker = {
            worker_index: annual_remaining[worker_index] / (13 - month)
            for worker_index in workers.index
        }

        solver = MonthlySolver(
            workers=workers,
            dates=dates,
            demands=demands,
            holidays=holidays,
            fixed_assignments=fixed_assignments,
            config=config,
            annual_hours_remaining=annual_remaining,
            months_remaining_including_current=(13 - month),
            previous_last_code=previous_last_code,
        )
        schedule, result = solver.solve()

        coverage = build_coverage(schedule, demands, holidays, dates)
        month_hours = build_month_hours(schedule, target_hours_by_worker, dates)

        for worker_index, worker_name in enumerate(workers["TRABAJADOR"]):
            assigned_hours = float(
                month_hours.loc[
                    month_hours["TRABAJADOR"] == worker_name,
                    "HORAS_ASIGNADAS_MES",
                ].iloc[0]
            )
            annual_remaining[worker_index] = max(
                0.0,
                annual_remaining[worker_index] - assigned_hours,
            )
            previous_last_code[worker_index] = schedule.iloc[worker_index][
                dates[-1].isoformat()
            ]

        schedule = schedule.drop(columns=["WORKER_IDX"])
        schedule.insert(0, "MES", month)
        coverage.insert(0, "MES", month)
        month_hours.insert(0, "MES", month)

        monthly_schedules[month] = schedule
        coverage_list.append(coverage)
        month_hours_list.append(month_hours)
        solve_log_rows.append(
            {
                "MES": month,
                "solver_status": getattr(result, "status", None),
                "solver_message": getattr(result, "message", None),
                "objective": float(getattr(result, "fun", np.nan)),
                "faltas_cobertura": int((~coverage["cumple"]).sum()),
                "desviacion_media_horas": float(
                    month_hours["DESVIACION_HORAS"].mean()
                ),
            }
        )

    coverage_all = pd.concat(coverage_list, ignore_index=True)
    monthly_hours_all = pd.concat(month_hours_list, ignore_index=True)
    annual_hours = (
        monthly_hours_all.groupby("TRABAJADOR", as_index=False)
        .agg(
            CATEGORIA=("CATEGORIA", "first"),
            **{
                "HORAS_ASIGNADAS_A\u00d1O": ("HORAS_ASIGNADAS_MES", "sum"),
                "DESCANSOS_A\u00d1O": ("DESCANSOS", "sum"),
            },
        )
        .merge(
            workers[["TRABAJADOR", "annual_hours"]].rename(
                columns={"annual_hours": "HORAS_OBJETIVO_A\u00d1O"}
            ),
            on="TRABAJADOR",
            how="left",
        )
    )
    annual_hours["DESVIACION_A\u00d1O"] = (
        annual_hours["HORAS_ASIGNADAS_A\u00d1O"]
        - annual_hours["HORAS_OBJETIVO_A\u00d1O"]
    )
    solve_log = pd.DataFrame(solve_log_rows)

    write_annual_workbook(
        monthly_schedules=monthly_schedules,
        coverage_all=coverage_all,
        monthly_hours_all=monthly_hours_all,
        annual_hours=annual_hours,
        solve_log=solve_log,
        config=config,
    )

    return AnnualRunArtifacts(
        config=config,
        workers=workers,
        demands=demands,
        holidays=holidays,
        fixed_assignments=fixed_assignments,
        monthly_schedules=monthly_schedules,
        coverage_all=coverage_all,
        monthly_hours_all=monthly_hours_all,
        annual_hours=annual_hours,
        solve_log=solve_log,
    )
